import os
import io
import zipfile
import csv
import datetime as dt
import xml.etree.ElementTree as ET
import unicodedata
import re
import tempfile
import logging
import math
from urllib.parse import urlencode
import zipstream
import json
import requests
from datetime import date
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.core.files import File
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import (
    Max,
    Min,
    F,
    Count,
    Q,
    Prefetch,
    Exists,
    OuterRef,
    Subquery,
    Case,
    When,
    BooleanField,
)
from django.http import (
    StreamingHttpResponse,
    FileResponse,
    JsonResponse,
    HttpResponseBadRequest,
    HttpResponse,
)
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.safestring import mark_safe
from django.views.decorators.http import require_GET, require_http_methods
from PIL import Image

from .forms import (
    WorkRecordForm,
    PhotoDocumentationForm,
    ProjectForm,
    CustomUserCreationForm,
    ProjectEditForm,
    AddMemberForm,
    TreeInterventionForm,
)
from .models import (
    Project,
    WorkRecord,
    PhotoDocumentation,
    ProjectMembership,
    TreeAssessment,
    ShrubAssessment,
    TreeIntervention,
    Species,
    MISTLETOE_LEVELS,
)
from .permissions import (
    user_projects_qs,
    user_can_view_project,
    user_is_foreman,
    get_project_or_404_for_user,
    is_project_member,
    can_edit_project,
    can_lock_project,
    can_delete_project,
    can_purge_project,
    can_transition_intervention,
)

# ------------------ Auth / základní stránky ------------------
logger = logging.getLogger(__name__)

def logout_view(request):
    logout(request)
    return redirect('login')

def home(request):
    return render(request, 'home.html')

def signup(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/signup.html', {'form': form})

# ------------------ Projekty ------------------

PROJECT_DETAIL_PAGE_SIZE = 40

def _project_detail_filters(request):
    q = request.GET.get('q', '').strip()
    df = parse_date(request.GET.get('date_from') or '')
    dt = parse_date(request.GET.get('date_to') or '')
    has_assessment = request.GET.get('has_assessment', '').strip()
    has_open_interventions = request.GET.get('has_open_interventions', '').strip()
    return q, df, dt, has_assessment, has_open_interventions

def _project_detail_queryset(project, q, df, dt, has_assessment, has_open_interventions):
    qs = project.trees.all()

    if q:
        qs = qs.filter(Q(title__icontains=q) | Q(description__icontains=q))
    if df:
        qs = qs.filter(date__gte=df)
    if dt:
        qs = qs.filter(date__lte=dt)

    if has_assessment in ("yes", "no"):
        tree_exists = Exists(TreeAssessment.objects.filter(work_record=OuterRef("pk")))
        shrub_exists = Exists(ShrubAssessment.objects.filter(work_record=OuterRef("pk")))
        has_assessment_expr = Case(
            When(
                vegetation_type__in=[
                    WorkRecord.VegetationType.SHRUB,
                    WorkRecord.VegetationType.HEDGE,
                ],
                then=shrub_exists,
            ),
            default=tree_exists,
            output_field=BooleanField(),
        )
        qs = qs.annotate(has_assessment_flag=has_assessment_expr)
        if has_assessment == "yes":
            qs = qs.filter(has_assessment_flag=True)
        else:
            qs = qs.filter(has_assessment_flag=False)

    if has_open_interventions == "none":
        qs = qs.filter(interventions__isnull=True)
    elif has_open_interventions == "proposed":
        qs = qs.filter(interventions__status="proposed").distinct()
    elif has_open_interventions == "approved":
        qs = qs.filter(interventions__status="done_pending_owner").distinct()
    elif has_open_interventions == "handover":
        qs = qs.filter(interventions__status="done_pending_owner").distinct()
    elif has_open_interventions == "completed":
        qs = qs.filter(interventions__status="completed").distinct()

    latest_assessments_qs = TreeAssessment.objects.order_by("-assessed_at", "-id")
    latest_shrub_assessments_qs = ShrubAssessment.objects.order_by("-assessed_at", "-id")
    interventions_qs = TreeIntervention.objects.select_related("intervention_type").order_by(
        "urgency", "due_date", "id"
    )

    return (
        qs.select_related("project")
        .prefetch_related(
            Prefetch("assessments", queryset=latest_assessments_qs, to_attr="prefetched_assessments"),
            Prefetch(
                "shrub_assessments",
                queryset=latest_shrub_assessments_qs,
                to_attr="prefetched_shrub_assessments",
            ),
            Prefetch("interventions", queryset=interventions_qs, to_attr="prefetched_interventions"),
        )
        .order_by("-created_at")
    )


def _decorate_interventions_for_user(user, interventions):
    for intervention in interventions:
        intervention.allowed_mark_done = can_transition_intervention(
            user, intervention, "done_pending_owner"
        )
        intervention.allowed_confirm = can_transition_intervention(
            user, intervention, "completed"
        )
        intervention.allowed_return = can_transition_intervention(
            user, intervention, "proposed"
        )
    return interventions


def _decorate_workrecords(work_records, user=None):
    for wr in work_records:
        interventions = list(getattr(wr, "prefetched_interventions", []))
        wr.interventions_list = interventions
        wr.intervention_count = len(interventions)
        wr.open_intervention_count = sum(1 for i in interventions if i.status != "completed")
        wr.max_urgency = max((i.urgency for i in interventions), default=None) if interventions else None

        wr.interventions_proposed = 0
        wr.interventions_approved = 0
        wr.interventions_handover = 0
        wr.interventions_completed = 0

        for iv in interventions:
            if iv.status == "proposed":
                wr.interventions_proposed += 1
            elif iv.status == "done_pending_owner":
                wr.interventions_handover += 1
            elif iv.status == "completed":
                wr.interventions_completed += 1

        if user and interventions:
            _decorate_interventions_for_user(user, interventions)


@login_required
def project_detail(request, pk):
    """Str nka vçech £kon… projektu + vyhled v n¡, filtry, str nkov n¡."""
    project = get_object_or_404(Project, pk=pk)

    # pr va
    if not user_can_view_project(request.user, project.pk):
        return redirect('work_record_list')

    q, df, dt, has_assessment, has_open_interventions = _project_detail_filters(request)
    qs = _project_detail_queryset(project, q, df, dt, has_assessment, has_open_interventions)

    paginator = Paginator(qs, PROJECT_DETAIL_PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get("page"))

    _decorate_workrecords(page_obj.object_list, request.user)

    has_next = page_obj.has_next()
    next_url = None
    if has_next:
        query = request.GET.copy()
        query["page"] = page_obj.next_page_number()
        next_url = f"{reverse('project_detail_items', args=[project.pk])}?{query.urlencode()}"

    # flagy pro çablonu (tlaŸ¡tka)
    is_member = is_project_member(request.user, project)
    can_edit = can_edit_project(request.user, project)
    can_lock = can_lock_project(request.user, project)
    can_delete = can_delete_project(request.user, project)

    return render(
        request,
        "tracker/project_detail.html",
        {
            "project": project,
            "page_obj": page_obj,
            "q": q,
            "date_from": request.GET.get("date_from", ""),
            "date_to": request.GET.get("date_to", ""),
            "has_assessment": has_assessment,
            "has_open_interventions": has_open_interventions,
            "can_edit_project": can_edit,
            "can_lock_project": can_lock,
            "can_delete_project": can_delete,
            "is_member": is_member,
            "has_next": has_next,
            "next_url": next_url,
            "project_return_url": request.get_full_path(),
        },
    )


@login_required
def project_detail_items(request, pk):
    project = get_object_or_404(Project, pk=pk)

    if not user_can_view_project(request.user, project.pk):
        return HttpResponse(status=403)

    q, df, dt, has_assessment, has_open_interventions = _project_detail_filters(request)
    qs = _project_detail_queryset(project, q, df, dt, has_assessment, has_open_interventions)

    paginator = Paginator(qs, PROJECT_DETAIL_PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get("page"))

    _decorate_workrecords(page_obj.object_list, request.user)

    response = render(
        request,
        "tracker/project_detail_items.html",
        {
            "work_records": page_obj.object_list,
            "project": project,
            "project_return_url": f"{reverse('project_detail', args=[project.pk])}?{request.GET.urlencode()}" if request.GET else reverse('project_detail', args=[project.pk]),
        },
    )

    response["X-Has-Next"] = "1" if page_obj.has_next() else "0"
    if page_obj.has_next():
        query = request.GET.copy()
        query["page"] = page_obj.next_page_number()
        response["X-Next-Url"] = f"{reverse('project_detail_items', args=[project.pk])}?{query.urlencode()}"

    return response


@login_required
def edit_project(request, pk):
    project = get_object_or_404(Project, pk=pk)

    # Jen stavbyvedoucí
    if not can_edit_project(request.user, project):
        return redirect('work_record_list')

    project_form = ProjectEditForm(instance=project)
    add_member_form = AddMemberForm()

    if request.method == 'POST':
        if 'save_project' in request.POST:
            project_form = ProjectEditForm(request.POST, instance=project)
            if project_form.is_valid():
                project_form.save()
                return redirect('work_record_list')

        elif 'add_member' in request.POST:
            add_member_form = AddMemberForm(request.POST)
            if add_member_form.is_valid():
                user = add_member_form.cleaned_data['user']
                role = add_member_form.cleaned_data['role']
                ProjectMembership.objects.get_or_create(
                    user=user,
                    project=project,
                    defaults={'role': role}
                )
                return redirect('edit_project', pk=pk)

    members = ProjectMembership.objects.filter(project=project).select_related('user')

    return render(request, 'tracker/edit_project.html', {
        'project': project,
        'project_form': project_form,
        'add_member_form': add_member_form,
        'members': members,
    })


@login_required
def remove_member(request, pk, user_id):
    project = get_object_or_404(Project, pk=pk)

    if not can_edit_project(request.user, project):
        return redirect('work_record_list')

    ProjectMembership.objects.filter(user_id=user_id, project=project).delete()
    return redirect('edit_project', pk=pk)


@login_required
def create_project(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.is_closed = False
            project.save()

            # autor = FOREMAN
            ProjectMembership.objects.get_or_create(
                user=request.user,
                project=project,
                defaults={"role": ProjectMembership.Role.FOREMAN},
            )

            return redirect('work_record_list')
    else:
        form = ProjectForm()
    return render(request, 'tracker/create_project.html', {'form': form})


@login_required
def close_project(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if not can_lock_project(request.user, project):
        return redirect('work_record_list')
    project.is_closed = True
    project.save()
    return redirect('work_record_list')


@login_required
def activate_project(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if not can_lock_project(request.user, project):
        return redirect('work_record_list')
    project.is_closed = False
    project.save()
    return redirect('closed_projects_list')

@login_required
@require_http_methods(["POST"])
def delete_project(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if not can_delete_project(request.user, project):
        return redirect('work_record_list')
    if not project.is_closed:
        messages.warning(request, "Projekt lze smazat pouze z uzavřených projektů.")
        return redirect('project_detail', pk=pk)
    project_name = project.name
    project.delete()
    messages.success(request, f"Projekt \"{project_name}\" byl smazán.")
    return redirect('closed_projects_list')

@login_required
@require_http_methods(["POST"])
def purge_project(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if not can_purge_project(request.user, project):
        return redirect('work_record_list')
    if not project.is_closed:
        messages.warning(request, "Projekt lze purgeovat pouze z uzavřených projektů.")
        return redirect('project_detail', pk=pk)

    confirm_text = (request.POST.get("confirm_text") or "").strip()
    if confirm_text not in (project.name, "DELETE"):
        messages.error(request, "Potvrzení nesouhlasí. Zadejte přesný název projektu nebo DELETE.")
        return redirect('closed_projects_list')

    work_records = WorkRecord.objects.filter(project=project).prefetch_related("photos")

    try:
        with transaction.atomic():
            for work_record in work_records:
                for photo in work_record.photos.all():
                    if not photo.photo:
                        continue
                    try:
                        photo.photo.delete(save=False)
                    except Exception:
                        logger.exception("Failed to delete photo file for PhotoDocumentation id=%s", photo.id)
                        raise RuntimeError("photo_delete_failed")

            work_records.delete()
            project.delete()
    except RuntimeError:
        messages.error(request, "Mazání souborů selhalo. Projekt nebyl smazán.")
        return redirect('closed_projects_list')

    messages.success(request, f"Projekt \"{project.name}\" byl purged včetně úkonů a fotek.")
    return redirect('closed_projects_list')


@login_required
def closed_projects_list(request):
    projects = (
        user_projects_qs(request.user)
        .filter(is_closed=True)
        .annotate(latest_work_time=Max('work_records__created_at'))
        .order_by(F('latest_work_time').desc(nulls_last=True), '-id')
    )

    # flag pro šablonu (zobrazit tlačítka jen foremanovi)
    for p in projects:
        p.is_foreman = can_edit_project(request.user, p)
        p.can_delete_project = can_delete_project(request.user, p)
        p.can_purge_project = can_purge_project(request.user, p)

    return render(request, 'tracker/closed_projects_list.html', {'projects': projects})

# ------------------ WorkRecord ------------------

@login_required
def work_record_list(request):
    latest_wr = WorkRecord.objects.order_by('-created_at')  # pro pořadí v prefetchi

    projects = (
        user_projects_qs(request.user)
        .filter(is_closed=False)
        .annotate(
            latest_work_time=Max('work_records__created_at'),
            trees_count=Count('trees', distinct=True)
        )
        .order_by(F('latest_work_time').desc(nulls_last=True), '-id')
        .prefetch_related(
            Prefetch('work_records', queryset=latest_wr)  # abychom měli nové nahoře
        )
    )

    for p in projects:
        p.is_foreman = can_edit_project(request.user, p)
        p.is_member = is_project_member(request.user, p)

    unassigned_count = None
    if request.user.is_superuser:
        unassigned_count = (
            WorkRecord.objects
            .filter(project__isnull=True, tree_projects__isnull=True)
            .distinct()
            .count()
        )

    return render(request, 'tracker/work_record_list.html', {
        'projects': projects,
        'unassigned_count': unassigned_count,
    })

@login_required
def unassigned_work_records_list(request):
    if not request.user.is_superuser:
        return redirect('work_record_list')

    qs = (
        WorkRecord.objects
        .filter(project__isnull=True, tree_projects__isnull=True)
        .order_by('-created_at')
    )
    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, 'tracker/unassigned_work_records_list.html', {
        'page_obj': page_obj,
    })

@login_required
def create_work_record(request, project_id=None):
    project_context_id = project_id or request.GET.get("project") or request.POST.get("project")
    if request.method == 'POST':
        work_record_form = WorkRecordForm(request.POST)
        photo_form = PhotoDocumentationForm(request.POST, request.FILES)

        if work_record_form.is_valid():
            work_record = work_record_form.save(commit=False)

            lat_str = request.POST.get("latitude") or request.GET.get("lat")
            lon_str = request.POST.get("longitude") or request.GET.get("lon")
            try:
                if lat_str is not None and lon_str is not None:
                    work_record.latitude = float(lat_str)
                    work_record.longitude = float(lon_str)
            except (TypeError, ValueError):
                pass

            work_record.save()
            work_record.assign_passport_identifiers()
            work_record.sync_title_from_identifiers()
            work_record.save(
                update_fields=[
                    "title",
                    "external_tree_id",
                    "description",
                    "date",
                    "project",
                    "latitude",
                    "longitude",
                ]
            )

            # ověř přístup k projektu
            if work_record.project_id and not user_can_view_project(request.user, work_record.project_id):
                return redirect('work_record_list')

            if project_context_id:
                try:
                    project = Project.objects.get(pk=project_context_id)
                except Project.DoesNotExist:
                    return redirect('work_record_list')
                if not user_can_view_project(request.user, project.pk):
                    return redirect('work_record_list')
                project.trees.add(work_record)

            if 'photo' in request.FILES and photo_form.is_valid():
                photo = photo_form.save(commit=False)
                photo.work_record = work_record
                photo.save()

            if project_context_id:
                return redirect('project_detail', pk=project_context_id)
            return redirect('work_record_detail', pk=work_record.pk)
    else:
        initial_data = {}
        if project_context_id:
            project = get_project_or_404_for_user(request.user, project_context_id)
            initial_data['project'] = project

        work_record_form = WorkRecordForm(initial=initial_data)
        photo_form = PhotoDocumentationForm()

    return render(request, 'tracker/create_work_record.html', {
        'work_record_form': work_record_form,
        'photo_form': photo_form,
        'project_context_id': project_context_id,
    })

@login_required
def delete_work_record(request, pk):
    """Smazání úkonu včetně fotek a souborů z disku."""
    work_record = get_object_or_404(WorkRecord, pk=pk)

    # kontrola oprávnění
    if not user_can_view_project(request.user, work_record.project.pk):
        return redirect('work_record_list')

    if request.method == "POST":
        # smažeme všechny fotky
        photos = PhotoDocumentation.objects.filter(work_record=work_record)
        for photo in photos:
            if photo.photo:
                # Nepouzivame .path; S3/django-storages nepodporuje absolutni cesty.
                try:
                    photo.photo.delete(save=False)
                except Exception as e:
                    print(f"⚠️ Nepodařilo se smazat soubor {photo.photo.name}: {e}")
            photo.delete()

        # smažeme samotný úkon
        work_record.delete()
        messages.success(request, "Úkon byl úspěšně smazán včetně fotek.")
        return redirect("project_detail", pk=work_record.project.pk)

    return render(request, "tracker/confirm_delete_work_record.html", {"work_record": work_record})

@login_required
def work_record_detail(request, pk):
    work_record = get_object_or_404(
        WorkRecord.objects
        .select_related("project")
        .prefetch_related("assessments", "shrub_assessments", "photos"),
        pk=pk,
    )

    if work_record.project_id and not user_can_view_project(request.user, work_record.project_id):
        return redirect('work_record_list')

    project_param = request.GET.get("project")
    return_url = request.GET.get("return")
    active_project_id = None
    if project_param:
        try:
            active_project_id = int(project_param)
        except (TypeError, ValueError):
            active_project_id = None
    if active_project_id is None:
        active_project_id = request.session.get("active_project_id")
    if not return_url:
        return_url = request.session.get("active_project_return")

    if request.method == 'POST':
        photo_form = PhotoDocumentationForm(request.POST, request.FILES)
        if photo_form.is_valid():
            photo = photo_form.save(commit=False)
            photo.work_record = work_record
            photo.save()
            return redirect('work_record_detail', pk=work_record.pk)
    else:
        photo_form = PhotoDocumentationForm()

    photos = work_record.photos.all()
    valid_photos = [photo for photo in photos if photo.photo]
    interventions = list(
        work_record.interventions.select_related("intervention_type").order_by(
            "status", "urgency", "due_date", "id"
        )
    )
    _decorate_interventions_for_user(request.user, interventions)
    current_interventions = [
        item for item in interventions if item.status in ("proposed", "done_pending_owner")
    ]
    history_interventions = [item for item in interventions if item.status == "completed"]

    return render(request, 'tracker/work_record_detail.html', {
        'work_record': work_record,
        'photo_form': photo_form,
        'photos': valid_photos,
        'interventions': interventions,
        'current_interventions': current_interventions,
        'history_interventions': history_interventions,
        'active_project_id': active_project_id,
        'return_url': return_url,
    })


@login_required
def tree_intervention_create(request, tree_id):
    tree = get_object_or_404(WorkRecord, pk=tree_id)
    if tree.project_id and not user_can_view_project(request.user, tree.project_id):
        return redirect('work_record_list')

    form = TreeInterventionForm(request.POST or None)
    intervention_note_data_json = mark_safe(json.dumps(form.intervention_type_note_data))

    if request.method == 'POST' and form.is_valid():
        intervention = form.save(commit=False)
        intervention.tree = tree
        intervention.created_by = request.user
        intervention.save()
        messages.success(request, "Zásah byl uložen.")
        return redirect('work_record_detail', pk=tree.pk)

    return render(request, 'tracker/tree_intervention_form.html', {
        'form': form,
        'tree': tree,
        'form_title': 'Přidat zásah',
        'intervention_note_data_json': intervention_note_data_json,
    })


@login_required
def tree_intervention_update(request, pk):
    intervention = get_object_or_404(TreeIntervention, pk=pk)
    tree = intervention.tree
    if tree.project_id and not user_can_view_project(request.user, tree.project_id):
        return redirect('work_record_list')

    form = TreeInterventionForm(request.POST or None, instance=intervention)
    intervention_note_data_json = mark_safe(json.dumps(form.intervention_type_note_data))

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Zásah byl aktualizován.")
        return redirect('work_record_detail', pk=tree.pk)

    return render(request, 'tracker/tree_intervention_form.html', {
        'form': form,
        'tree': tree,
        'intervention': intervention,
        'form_title': 'Upravit zásah',
        'intervention_note_data_json': intervention_note_data_json,
    })


@login_required
@require_http_methods(["POST"])
def tree_intervention_transition(request, pk):
    intervention = get_object_or_404(
        TreeIntervention.objects.select_related("tree__project"),
        pk=pk,
    )
    tree = intervention.tree
    if tree.project_id and not user_can_view_project(request.user, tree.project_id):
        return redirect("work_record_list")

    target = (request.POST.get("target") or request.POST.get("action") or "").strip()
    note = (request.POST.get("note") or "").strip()

    if target not in ("proposed", "done_pending_owner", "completed"):
        return HttpResponseBadRequest("Invalid target")

    requires_note = target == "proposed" and intervention.status in ("done_pending_owner", "completed")
    if requires_note and not note:
        messages.error(request, "Poznámka je povinná.")
        return redirect(request.META.get("HTTP_REFERER") or reverse("work_record_detail", args=[tree.pk]))

    if not can_transition_intervention(request.user, intervention, target):
        return HttpResponse(status=403)

    intervention.status = target
    update_fields = ["status"]
    if note:
        intervention.status_note = note
        update_fields.append("status_note")
    if target == "done_pending_owner" and getattr(intervention, "handed_over_for_check_at", None) is None:
        intervention.handed_over_for_check_at = timezone.now()
        update_fields.append("handed_over_for_check_at")
    if target == "completed" and getattr(intervention, "approved_at", None) is None:
        intervention.approved_at = timezone.now()
        update_fields.append("approved_at")
    intervention.save(update_fields=update_fields)

    return redirect(request.META.get("HTTP_REFERER") or reverse("work_record_detail", args=[tree.pk]))


@login_required
@require_http_methods(["POST"])
def tree_intervention_api(request, tree_id):
    tree = get_object_or_404(WorkRecord, pk=tree_id)
    if tree.project_id and not user_can_view_project(request.user, tree.project_id):
        return JsonResponse({'status': 'error', 'msg': 'Nemáte oprávnění.'}, status=403)

    form = TreeInterventionForm(request.POST)
    if not form.is_valid():
        return JsonResponse({'status': 'error', 'errors': form.errors.get_json_data()}, status=400)

    intervention = form.save(commit=False)
    intervention.tree = tree
    intervention.created_by = request.user
    intervention.save()

    return JsonResponse({
        'status': 'ok',
        'intervention': {
            'id': intervention.pk,
            'type': str(intervention.intervention_type),
            'code': intervention.intervention_type.code,
            'urgency': intervention.get_urgency_display(),
            'status': intervention.get_status_display(),
        },
    })


@login_required
@require_http_methods(["GET", "POST"])
def tree_intervention_api(request, tree_id):
    tree = get_object_or_404(WorkRecord, pk=tree_id)
    if tree.project_id and not user_can_view_project(request.user, tree.project_id):
        return JsonResponse({'status': 'error', 'msg': 'Nemáte oprávnění.'}, status=403)

    def serialize_intervention(obj):
        allowed_actions = {
            "mark_done": can_transition_intervention(request.user, obj, "done_pending_owner"),
            "confirm": can_transition_intervention(request.user, obj, "completed"),
            "return": can_transition_intervention(request.user, obj, "proposed"),
        }
        return {
            'id': obj.pk,
            'code': obj.intervention_type.code if obj.intervention_type else '',
            'name': obj.intervention_type.name if obj.intervention_type else '',
            'urgency': obj.get_urgency_display(),
            'status': obj.get_status_display(),
            'status_code': obj.status,
            'status_note': obj.status_note or '',
            'description': obj.description or '',
            'transition_url': reverse('tree_intervention_transition', args=[obj.pk]),
            'created_at': obj.created_at.isoformat() if obj.created_at else None,
            'handed_over_for_check_at': (
                obj.handed_over_for_check_at.isoformat()
                if getattr(obj, "handed_over_for_check_at", None)
                else None
            ),
            'allowed_actions': allowed_actions,
        }

    if request.method == "GET":
        interventions = (
            TreeIntervention.objects
            .filter(tree=tree)
            .select_related("intervention_type")
            .order_by("status", "urgency", "due_date", "id")
        )
        data = [serialize_intervention(obj) for obj in interventions]
        return JsonResponse({'status': 'ok', 'interventions': data})

    action = request.POST.get('action')
    if action == "handover":
        try:
            intervention_id = int(request.POST.get("id") or 0)
        except (TypeError, ValueError):
            return JsonResponse({'status': 'error', 'msg': 'Neplatné ID zásahu.'}, status=400)

        intervention = get_object_or_404(
            TreeIntervention.objects.select_related("intervention_type"),
            pk=intervention_id,
            tree=tree,
        )
        if intervention.status != "proposed":
            return JsonResponse({'status': 'error', 'msg': 'Tento zásah nelze předat ke kontrole.'}, status=400)
        if not can_transition_intervention(request.user, intervention, "done_pending_owner"):
            return JsonResponse({'status': 'error', 'msg': 'Nemáte oprávnění.'}, status=403)

        intervention.status = "done_pending_owner"
        if getattr(intervention, "handed_over_for_check_at", None) is None:
            intervention.handed_over_for_check_at = timezone.now()
        intervention.save(update_fields=["status", "handed_over_for_check_at"])
        return JsonResponse({'status': 'ok', 'intervention': serialize_intervention(intervention)})

    form = TreeInterventionForm(request.POST)
    if not form.is_valid():
        return JsonResponse({'status': 'error', 'errors': form.errors.get_json_data()}, status=400)

    intervention = form.save(commit=False)
    intervention.tree = tree
    intervention.created_by = request.user
    intervention.save()

    return JsonResponse({'status': 'ok', 'intervention': serialize_intervention(intervention)})


@login_required
def edit_work_record(request, pk):
    work_record = get_object_or_404(WorkRecord, pk=pk)

    # přístup jen pro členy projektu
    if work_record.project_id and not user_can_view_project(request.user, work_record.project_id):
        return redirect('work_record_list')

    work_record_form = WorkRecordForm(instance=work_record)
    photo_form = PhotoDocumentationForm()

    if request.method == 'POST':
        if 'save_work_record' in request.POST:
            work_record_form = WorkRecordForm(request.POST, instance=work_record)
            if work_record_form.is_valid():
                work_record = work_record_form.save(commit=False)
                work_record.sync_title_from_identifiers()
                work_record.save()
                from django.urls import reverse
                url = reverse('work_record_list')
                if work_record.project_id:
                    url = f"{url}#project-{work_record.project_id}"
                return redirect(url)

        elif 'add_photo' in request.POST:
            photo_form = PhotoDocumentationForm(request.POST, request.FILES)
            if 'photo' in request.FILES and photo_form.is_valid():
                photo = photo_form.save(commit=False)
                photo.work_record = work_record
                photo.save()
            return redirect('edit_work_record', pk=work_record.pk)

    photos = work_record.photos.all()

    return render(request, 'tracker/edit_work_record.html', {
        'work_record_form': work_record_form,
        'photo_form': photo_form,
        'photos': photos,
        'work_record': work_record,
    })

# ------------------ Fotky ------------------

@login_required
def add_photo(request, work_record_id):
    work_record = get_object_or_404(WorkRecord, pk=work_record_id)

    if work_record.project_id and not user_can_view_project(request.user, work_record.project_id):
        return redirect('work_record_list')

    if request.method == "POST":
        form = PhotoDocumentationForm(request.POST, request.FILES)
        photo = request.FILES.get("photo")

        if not photo:
            default_photo_path = os.path.join(settings.MEDIA_ROOT, "photos", "default.jpg")
            with open(default_photo_path, 'rb') as default_file:
                photo = File(default_file, name="default.jpg")

        if form.is_valid():
            new_photo = form.save(commit=False)
            new_photo.work_record = work_record
            new_photo.photo = photo
            new_photo.save()

    return redirect("work_record_detail", pk=work_record_id)


@login_required
def delete_photo(request, pk):
    photo = get_object_or_404(PhotoDocumentation, pk=pk)

    if photo.work_record.project_id and not user_can_view_project(request.user, photo.work_record.project_id):
        return redirect('work_record_list')

    work_record_pk = photo.work_record.pk
    photo.delete()
    return redirect('edit_work_record', pk=work_record_pk)




def _slugify_export_name(name):
    name = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')
    name = re.sub(r'[^a-zA-Z0-9_-]+', '_', name)
    return name.strip('_') or 'ukon'


def _get_export_work_records(request, project):
    export_all_requested = bool(request.POST.get("export_all"))
    if export_all_requested:
        return project.trees.all(), None

    selected_ids = request.POST.getlist("selected_records")
    if not selected_ids:
        messages.warning(request, "Vyberte prosim alespon jeden zaznam pro export.")
        return None, redirect("project_detail", pk=project.pk)

    return project.trees.filter(id__in=selected_ids), None


def _build_export_queryset(work_records):
    latest_assessments_qs = TreeAssessment.objects.order_by("-assessed_at", "-id")
    latest_shrub_assessments_qs = ShrubAssessment.objects.order_by("-assessed_at", "-id")
    return (
        work_records.select_related("project")
        .prefetch_related(
            Prefetch(
                "assessments",
                queryset=latest_assessments_qs,
                to_attr="prefetched_assessments",
            ),
            Prefetch(
                "shrub_assessments",
                queryset=latest_shrub_assessments_qs,
                to_attr="prefetched_shrub_assessments",
            ),
            "interventions__intervention_type",
        )
        .annotate(intervention_count=Count("interventions", distinct=True))
    )


def _latest_assessment_for_export(record):
    assessments = getattr(record, "prefetched_assessments", None)
    if assessments:
        return assessments[0]
    return None


def _latest_shrub_assessment_for_export(record):
    assessments = getattr(record, "prefetched_shrub_assessments", None)
    if assessments:
        return assessments[0]
    return None


@login_required
def export_selected_zip(request, pk):
    """Export vybraných nebo všech úkonů projektu jako streamovaný ZIP (nízká paměťová náročnost)."""
    project = get_object_or_404(Project, pk=pk)

    if not user_can_view_project(request.user, project.pk):
        return redirect('work_record_list')

    work_records, redirect_response = _get_export_work_records(request, project)
    if redirect_response:
        return redirect_response


    def file_chunks(file_obj):
        if hasattr(file_obj, "chunks"):
            for chunk in file_obj.chunks():
                if chunk:
                    yield chunk
        else:
            while True:
                chunk = file_obj.read(64 * 1024)
                if not chunk:
                    break
                yield chunk

    # vytvoření streamovacího ZIP objektu
    z = zipstream.ZipFile(mode='w', compression=zipfile.ZIP_DEFLATED)

    used_folders = set()
    for record in work_records:
        base_label = record.preferred_id_label or f"ukon_{record.id}"
        folder = _slugify_export_name(base_label) or f"ukon_{record.id}"
        if folder in used_folders:
            folder = f"{folder}_{record.id}"
        used_folders.add(folder)
        photos = PhotoDocumentation.objects.filter(work_record=record)

        for photo in photos:
            if not photo.photo:
                continue

            base_filename = os.path.basename(photo.photo.name)
            ext = os.path.splitext(base_filename)[1] or ".jpg"

            if photo.description:
                safe_name = _slugify_export_name(photo.description)
                filename = f"{safe_name}{ext}"
            else:
                filename = base_filename

            arcname = os.path.join(folder, filename)
            storage = photo.photo.storage
            name = photo.photo.name
            local_path = None
            try:
                local_path = storage.path(name)
            except (NotImplementedError, AttributeError):
                local_path = None

            if local_path and os.path.exists(local_path):
                z.write(local_path, arcname)
                continue

            try:
                file_handle = storage.open(name, "rb")
            except Exception:
                continue

            with file_handle as f:
                z.write_iter(arcname, file_chunks(f))

    # příprava odpovědi
    today_str = date.today().strftime("%Y-%m-%d")
    filename = f'{_slugify_export_name(project.name)}_{today_str}.zip'

    response = StreamingHttpResponse(z, content_type='application/zip')
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

EXPORT_HEADERS = [
    "work_record_id",
    "project_id",
    "project_name",
    "title",
    "external_tree_id",
    "passport_no",
    "passport_code",
    "vegetation_type",
    "taxon",
    "taxon_czech",
    "taxon_latin",
    "latitude",
    "longitude",
    "date",
    "created_at",
    "parcel_number",
    "cadastral_area_code",
    "cadastral_area_name",
    "municipality_code",
    "municipality_name",
    "lv_number",
    "cad_lookup_status",
    "cad_lookup_at",
    "intervention_count",
    "interventions_codes",
    "assessment_assessed_at",
    "assessment_dbh_cm",
    "assessment_stem_circumference_cm",
    "assessment_stem_diameters_cm_list",
    "assessment_stem_circumferences_cm_list",
    "assessment_height_m",
    "assessment_crown_width_m",
    "assessment_crown_area_m2",
    "assessment_physiological_age",
    "assessment_vitality",
    "assessment_health_state",
    "assessment_stability",
    "assessment_mistletoe_level_raw",
    "assessment_mistletoe_text",
    "assessment_perspective",
    "shrub_assessed_at",
    "shrub_vitality",
    "shrub_height_m",
    "shrub_width_m",
    "shrub_note",
]


def _format_csv_list(value):
    if not value:
        return ""
    parts = [part.strip() for part in str(value).split(",")]
    parts = [part for part in parts if part]
    return ", ".join(parts)


def _mistletoe_text(level):
    if not level:
        return "bez jmelí"
    try:
        key = int(level)
    except (TypeError, ValueError):
        return ""
    info = MISTLETOE_LEVELS.get(key)
    if not info:
        return ""
    return f"{info['code']} – {info['label']} ({info['range']} objemu koruny)"


def _interventions_codes(record):
    codes = set()
    for intervention in record.interventions.all():
        if intervention.intervention_type and intervention.intervention_type.code:
            codes.add(intervention.intervention_type.code)
    return ", ".join(sorted(codes))


def _export_row_native(record, assessment, shrub_assessment):
    def to_float(value):
        if value is None:
            return None
        return float(value)

    shrub_kind = record.vegetation_type in (
        WorkRecord.VegetationType.SHRUB,
        WorkRecord.VegetationType.HEDGE,
    )
    return [
        record.id,
        record.project_id,
        record.project.name if record.project else None,
        record.title or None,
        record.external_tree_id or None,
        record.passport_no,
        record.passport_code or None,
        record.vegetation_type,
        record.taxon or None,
        record.taxon_czech or None,
        record.taxon_latin or None,
        to_float(record.latitude) if record.latitude is not None else None,
        to_float(record.longitude) if record.longitude is not None else None,
        record.date if record.date else None,
        record.created_at if record.created_at else None,
        record.parcel_number,
        record.cadastral_area_code,
        record.cadastral_area_name,
        record.municipality_code,
        record.municipality_name,
        record.lv_number,
        record.cad_lookup_status,
        record.cad_lookup_at if record.cad_lookup_at else None,
        getattr(record, "intervention_count", None),
        _interventions_codes(record),
        assessment.assessed_at if assessment and assessment.assessed_at else None,
        to_float(assessment.dbh_cm) if assessment and assessment.dbh_cm is not None else None,
        to_float(assessment.stem_circumference_cm)
        if assessment and assessment.stem_circumference_cm is not None
        else None,
        _format_csv_list(assessment.stem_diameters_cm_list) if assessment else "",
        _format_csv_list(assessment.stem_circumferences_cm_list) if assessment else "",
        to_float(assessment.height_m) if assessment and assessment.height_m is not None else None,
        to_float(assessment.crown_width_m) if assessment and assessment.crown_width_m is not None else None,
        to_float(assessment.crown_area_m2) if assessment and assessment.crown_area_m2 is not None else None,
        assessment.physiological_age if assessment and assessment.physiological_age is not None else None,
        assessment.vitality if assessment and assessment.vitality is not None else None,
        assessment.health_state if assessment and assessment.health_state is not None else None,
        assessment.stability if assessment and assessment.stability is not None else None,
        assessment.mistletoe_level if assessment and assessment.mistletoe_level is not None else None,
        _mistletoe_text(assessment.mistletoe_level) if assessment else "",
        assessment.perspective if assessment and assessment.perspective is not None else None,
        shrub_assessment.assessed_at
        if shrub_kind and shrub_assessment and shrub_assessment.assessed_at
        else None,
        shrub_assessment.vitality
        if shrub_kind and shrub_assessment and shrub_assessment.vitality is not None
        else None,
        to_float(shrub_assessment.height_m)
        if shrub_kind and shrub_assessment and shrub_assessment.height_m is not None
        else None,
        to_float(shrub_assessment.width_m)
        if shrub_kind and shrub_assessment and shrub_assessment.width_m is not None
        else None,
        shrub_assessment.note if shrub_kind and shrub_assessment else "",
    ]




@login_required
def export_selected_csv(request, pk):
    project = get_object_or_404(Project, pk=pk)

    if not user_can_view_project(request.user, project.pk):
        return redirect('work_record_list')

    work_records, redirect_response = _get_export_work_records(request, project)
    if redirect_response:
        return redirect_response

    work_records = _build_export_queryset(work_records)

    today_str = date.today().strftime("%Y-%m-%d")
    filename = f'{_slugify_export_name(project.name)}_{today_str}.csv'

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(EXPORT_HEADERS)

    for record in work_records:
        assessment = _latest_assessment_for_export(record)
        shrub_assessment = _latest_shrub_assessment_for_export(record)
        row = _export_row_native(record, assessment, shrub_assessment)
        csv_row = []
        for value in row:
            if isinstance(value, (dt.datetime, date)):
                csv_row.append(value.isoformat())
            elif value is None:
                csv_row.append("")
            else:
                csv_row.append(value)
        writer.writerow(
            csv_row
        )

    return response





@login_required
def export_selected_xlsx(request, pk):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        messages.error(
            request,
            "Export Excelu vyzaduje balicek openpyxl. Nainstalujte jej prosim.",
        )
        return redirect("project_detail", pk=pk)

    def excel_safe(value):
        if isinstance(value, dt.datetime):
            if timezone.is_aware(value):
                return timezone.localtime(value).replace(tzinfo=None)
            return value
        return value

    project = get_object_or_404(Project, pk=pk)

    if not user_can_view_project(request.user, project.pk):
        return redirect('work_record_list')

    work_records, redirect_response = _get_export_work_records(request, project)
    if redirect_response:
        return redirect_response

    work_records = _build_export_queryset(work_records)

    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(EXPORT_HEADERS)

    for record in work_records:
        assessment = _latest_assessment_for_export(record)
        shrub_assessment = _latest_shrub_assessment_for_export(record)
        row = _export_row_native(record, assessment, shrub_assessment)
        ws.append([excel_safe(item) for item in row])

    ws.freeze_panes = "A2"
    last_row = ws.max_row
    last_col_letter = get_column_letter(len(EXPORT_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"arbomap_project_{project.pk}_data.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_selected_xml(request, pk):
    project = get_object_or_404(Project, pk=pk)

    if not user_can_view_project(request.user, project.pk):
        return redirect('work_record_list')

    work_records, redirect_response = _get_export_work_records(request, project)
    if redirect_response:
        return redirect_response

    work_records = _build_export_queryset(work_records)

    root = ET.Element("arbomap_export", generated_at=timezone.now().isoformat())
    project_el = ET.SubElement(
        root,
        "project",
        id=str(project.pk),
        name=project.name or "",
    )

    for record in work_records:
        wr_el = ET.SubElement(project_el, "work_record", id=str(record.pk))
        loc_attrs = {}
        if record.latitude is not None:
            loc_attrs["lat"] = str(record.latitude)
        if record.longitude is not None:
            loc_attrs["lon"] = str(record.longitude)
        if loc_attrs:
            ET.SubElement(wr_el, "location", **loc_attrs)

        assessment = _latest_assessment_for_export(record)
        if assessment:
            attrs = {}
            if assessment.height_m is not None:
                attrs["height_m"] = str(assessment.height_m)
            if assessment.crown_width_m is not None:
                attrs["crown_width_m"] = str(assessment.crown_width_m)
            if assessment.crown_area_m2 is not None:
                attrs["crown_area_m2"] = str(assessment.crown_area_m2)
            if assessment.dbh_cm is not None:
                attrs["dbh_cm"] = str(assessment.dbh_cm)
            if assessment.physiological_age is not None:
                attrs["physiological_age"] = str(assessment.physiological_age)
            if assessment.vitality is not None:
                attrs["vitality"] = str(assessment.vitality)
            if assessment.health_state is not None:
                attrs["health_state"] = str(assessment.health_state)
            if assessment.stability is not None:
                attrs["stability"] = str(assessment.stability)
            if assessment.perspective:
                attrs["perspective"] = assessment.perspective
            if assessment.assessed_at:
                attrs["assessed_at"] = assessment.assessed_at.isoformat()
            ET.SubElement(wr_el, "assessment", **attrs)

    ET.indent(root, space="  ")
    xml_bytes = ET.tostring(root, encoding="utf-8", xml_declaration=True) + b"\n"

    filename = f"arbomap_project_{project.pk}.xml"
    response = HttpResponse(xml_bytes, content_type='application/xml; charset=utf-8')
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_qgis_geojson(request, pk):
    project = get_object_or_404(Project, pk=pk)

    if not user_can_view_project(request.user, project.pk):
        return redirect('work_record_list')

    work_records, redirect_response = _get_export_work_records(request, project)
    if redirect_response:
        return redirect_response

    work_records = _build_export_queryset(work_records)

    tree_features = []
    hedge_features = []

    for record in work_records:
        assessment = _latest_assessment_for_export(record)
        shrub_assessment = _latest_shrub_assessment_for_export(record)

        properties = {
            "work_record_id": record.id,
            "project_id": record.project_id,
            "project_name": record.project.name if record.project else None,
            "title": record.title or None,
            "external_tree_id": record.external_tree_id or None,
            "passport_no": record.passport_no,
            "passport_code": record.passport_code or None,
            "vegetation_type": record.vegetation_type,
            "taxon": record.taxon or None,
            "taxon_czech": record.taxon_czech or None,
            "taxon_latin": record.taxon_latin or None,
            "latitude": record.latitude,
            "longitude": record.longitude,
            "date": record.date.isoformat() if record.date else None,
            "created_at": record.created_at.isoformat() if record.created_at else None,
            "parcel_number": record.parcel_number,
            "cadastral_area_code": record.cadastral_area_code,
            "cadastral_area_name": record.cadastral_area_name,
            "municipality_code": record.municipality_code,
            "municipality_name": record.municipality_name,
            "lv_number": record.lv_number,
            "cad_lookup_status": record.cad_lookup_status,
            "cad_lookup_at": record.cad_lookup_at.isoformat() if record.cad_lookup_at else None,
            "intervention_count": getattr(record, "intervention_count", None),
            "interventions_codes": _interventions_codes(record),
            "assessment_assessed_at": assessment.assessed_at.isoformat()
            if assessment and assessment.assessed_at
            else None,
            "assessment_dbh_cm": assessment.dbh_cm if assessment else None,
            "assessment_stem_circumference_cm": assessment.stem_circumference_cm if assessment else None,
            "assessment_stem_diameters_cm_list": _format_csv_list(assessment.stem_diameters_cm_list)
            if assessment
            else "",
            "assessment_stem_circumferences_cm_list": _format_csv_list(
                assessment.stem_circumferences_cm_list
            )
            if assessment
            else "",
            "assessment_height_m": assessment.height_m if assessment else None,
            "assessment_crown_width_m": assessment.crown_width_m if assessment else None,
            "assessment_crown_area_m2": assessment.crown_area_m2 if assessment else None,
            "assessment_physiological_age": assessment.physiological_age if assessment else None,
            "assessment_vitality": assessment.vitality if assessment else None,
            "assessment_health_state": assessment.health_state if assessment else None,
            "assessment_stability": assessment.stability if assessment else None,
            "assessment_mistletoe_level_raw": assessment.mistletoe_level if assessment else None,
            "assessment_mistletoe_text": _mistletoe_text(
                assessment.mistletoe_level if assessment else None
            ),
            "assessment_perspective": assessment.perspective if assessment else None,
            "shrub_assessed_at": shrub_assessment.assessed_at.isoformat()
            if shrub_assessment and shrub_assessment.assessed_at
            else None,
            "shrub_vitality": shrub_assessment.vitality if shrub_assessment else None,
            "shrub_height_m": shrub_assessment.height_m if shrub_assessment else None,
            "shrub_width_m": shrub_assessment.width_m if shrub_assessment else None,
            "shrub_note": shrub_assessment.note if shrub_assessment else "",
        }

        if record.vegetation_type == WorkRecord.VegetationType.HEDGE:
            if not record.hedge_line:
                continue
            hedge_features.append(
                {
                    "type": "Feature",
                    "geometry": record.hedge_line,
                    "properties": properties,
                }
            )
            continue

        if record.latitude is None or record.longitude is None:
            continue
        tree_features.append(
            {
                "type": "Feature",
                "geometry": {
                    "type": "Point",
                    "coordinates": [record.longitude, record.latitude],
                },
                "properties": properties,
            }
        )

    def _json_safe(value):
        if isinstance(value, dt.datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, dict):
            return {key: _json_safe(val) for key, val in value.items()}
        if isinstance(value, (list, tuple)):
            return [_json_safe(item) for item in value]
        return value

    trees_geojson = _json_safe({"type": "FeatureCollection", "features": tree_features})
    hedges_geojson = _json_safe({"type": "FeatureCollection", "features": hedge_features})

    output = io.BytesIO()
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("trees_points.geojson", json.dumps(trees_geojson, ensure_ascii=False))
        zf.writestr("hedges_lines.geojson", json.dumps(hedges_geojson, ensure_ascii=False))

    output.seek(0)
    today_str = date.today().strftime("%Y-%m-%d")
    filename = f'{_slugify_export_name(project.name)}_{today_str}_qgis_geojson.zip'
    response = HttpResponse(output.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


"""
# ------------------ Test mapy ------------------

def map_test(request):
    # testovací souřadnice – Třinec
    context = {
        "latitude": 49.684,
        "longitude": 18.676,
        "mapy_key": settings.MAPY_API_KEY,
    }
    return render(request, "tracker/map_test.html", context)

def mapy_key_test(request):
    return JsonResponse({
        "mapy_key": settings.MAPY_API_KEY or "❌ Žádný klíč nenalezen"
    })

def mapy_geocode_test(request):
    query = request.GET.get("q", "Ostrava")
    url = "https://api.mapy.cz/v1/geocode"
    params = {"query": query, "lang": "cs", "key": settings.MAPY_API_KEY}

    r = requests.get(url, params=params)
    data = r.json()
    print(data)

    return render(request, "tracker/map_rest_test.html", {"data": data, "query": query})


"""

THUMB_MAX_SIZE = 256


@login_required
def bulk_approve_interventions(request, pk):
    project = get_object_or_404(Project, pk=pk)

    if not user_can_view_project(request.user, project.pk):
        return redirect('work_record_list')

    if request.method != "POST":
        return redirect("project_detail", pk=pk)

    selected_ids = request.POST.getlist("selected_records")
    if not selected_ids:
        messages.warning(request, "Vyberte prosím alespoň jeden strom / úkon.")
        return redirect("project_detail", pk=pk)

    work_records = project.trees.filter(id__in=selected_ids)
    interventions_qs = TreeIntervention.objects.filter(
        tree__in=work_records,
        status__in=["proposed"],
    )
    interventions = list(interventions_qs)

    if not interventions:
        messages.info(request, "Nebyl nalezen žádný navržený zásah ke schválení.")
        return redirect("project_detail", pk=pk)

    for intervention in interventions:
        if not can_transition_intervention(request.user, intervention, "done_pending_owner"):
            return HttpResponse(status=403)
        intervention.status = "done_pending_owner"
        if getattr(intervention, "approved_at", None) is None:
            intervention.approved_at = timezone.now()
        intervention.save(update_fields=["status", "approved_at"])
    messages.success(
        request,
        f"Schváleno {len(interventions)} navržených zásahů.",
    )
    return redirect("project_detail", pk=pk)


def get_photo_thumbnail(photo_obj, size=THUMB_MAX_SIZE):
    """
    Returns URL of a cached thumbnail for the given photo. Generates it on demand.
    """
    if not photo_obj or not photo_obj.photo:
        return None

    storage = photo_obj.photo.storage
    thumb_rel_path = f"photos/thumbs/{photo_obj.id}_{size}.jpg"
    if storage.exists(thumb_rel_path):
        return storage.url(thumb_rel_path)

    try:
        photo_obj.photo.open()
        with Image.open(photo_obj.photo) as img:
            img = img.convert("RGB")
            img.thumbnail((size, size))
            buffer = io.BytesIO()
            img.save(buffer, format="JPEG", quality=70, optimize=True)
        buffer.seek(0)
        storage.save(thumb_rel_path, ContentFile(buffer.getvalue()))
        return storage.url(thumb_rel_path)
    except Exception:
        return photo_obj.photo.url


def _build_map_mapui_context(request):
    visible_projects = user_projects_qs(request.user)
    base_records = (
        WorkRecord.objects
        .filter(Q(project__in=visible_projects) | Q(project__isnull=True))
        .select_related("project")
    )
    records = list(base_records.order_by("-id"))
    tree_assessment_exists = Exists(TreeAssessment.objects.filter(work_record=OuterRef("pk")))
    shrub_assessment_exists = Exists(ShrubAssessment.objects.filter(work_record=OuterRef("pk")))
    coords_qs = (
        base_records
        .filter(latitude__isnull=False, longitude__isnull=False)
        .annotate(
            photo_count=Count("photos", distinct=True),
            has_any_assessment=Case(
                When(
                    vegetation_type__in=[
                        WorkRecord.VegetationType.SHRUB,
                        WorkRecord.VegetationType.HEDGE,
                    ],
                    then=shrub_assessment_exists,
                ),
                default=tree_assessment_exists,
                output_field=BooleanField(),
            ),
        )
        .order_by("-id")
    )
    projects = visible_projects.order_by("name")
    projects_js = list(projects.values("id", "name"))
    coords = []
    for r in coords_qs:
        coords.append({
            "id": r.id,
            "title": r.title or "",
            "external_tree_id": r.external_tree_id or "",
            "passport_code": r.passport_code or "",
            "display_label": r.display_label,
            "taxon": r.taxon or "",
            "project": r.project.name if r.project else "",
            "project_id": r.project_id,
            "lat": r.latitude,
            "lon": r.longitude,
            "vegetation_type": r.vegetation_type,
            "has_assessment": bool(getattr(r, "has_any_assessment", False)),
            "has_photos": bool(getattr(r, "photo_count", 0)),
        })

    records_for_select = [
        {
            "id": r.id,
            "title": (r.title or ""),
            "taxon": r.taxon or "",
            "external_tree_id": r.external_tree_id or "",
            "passport_code": r.passport_code or "",
            "display_label": r.display_label,
            "project_id": r.project_id,
            "has_coords": (r.latitude is not None and r.longitude is not None),
        }
        for r in records
    ]

    intervention_form = TreeInterventionForm()
    intervention_note_data_json = mark_safe(
        json.dumps(intervention_form.intervention_type_note_data)
    )

    return {
        "mapy_key": settings.MAPY_API_KEY,
        "projects": projects,
        "projects_js": projects_js,
        "work_records": records,
        "records_with_coords": coords,
        "records_for_select": records_for_select,
        "intervention_form": intervention_form,
        "intervention_note_data_json": intervention_note_data_json,
    }


@login_required
def map_leaflet_test(request):
    target = reverse("map_gl_pilot")
    query = request.META.get("QUERY_STRING")
    if query:
        return redirect(f"{target}?{query}")
    return redirect(target)


@login_required
def workrecords_geojson(request):
    """
    GeoJSON feed with coordinates of WorkRecords (pilot usage for MapLibre map).
    """
    def to_float(value):
        if value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    project_param = request.GET.get("project")
    if project_param:
        try:
            project_id = int(project_param)
        except (TypeError, ValueError):
            return HttpResponseBadRequest("Invalid project parameter")
        project = get_object_or_404(Project, pk=project_id)
        if not user_can_view_project(request.user, project.pk):
            return JsonResponse({"error": "Forbidden"}, status=403)
        qs = project.trees.filter(
            latitude__isnull=False,
            longitude__isnull=False,
        ).only(
            "id",
            "latitude",
            "longitude",
            "external_tree_id",
            "title",
            "passport_code",
            "vegetation_type",
            "hedge_line",
        )
    else:
        # Use the Project.trees M2M as the source of truth; legacy FK can drift.
        visible_projects = user_projects_qs(request.user)
        qs = (
            WorkRecord.objects.filter(
                Q(projects__in=visible_projects),
                latitude__isnull=False,
                longitude__isnull=False,
            )
            .distinct()
            .only(
                "id",
                "latitude",
                "longitude",
                "external_tree_id",
                "title",
                "passport_code",
                "vegetation_type",
                "hedge_line",
            )
        )

    latest_assessment = TreeAssessment.objects.filter(work_record=OuterRef("pk")).order_by(
        "-assessed_at",
        "-id",
    )
    latest_shrub = ShrubAssessment.objects.filter(work_record=OuterRef("pk")).order_by(
        "-assessed_at",
        "-id",
    )
    approved_interventions = TreeIntervention.objects.filter(
        tree=OuterRef("pk"),
        status="completed",
    )
    done_interventions = TreeIntervention.objects.filter(
        tree=OuterRef("pk"),
        status="done_pending_owner",
    )
    qs = qs.annotate(
        crown_width_m=Subquery(latest_assessment.values("crown_width_m")[:1]),
        shrub_width_m=Subquery(latest_shrub.values("width_m")[:1]),
        has_approved_intervention=Exists(approved_interventions),
        has_done_intervention=Exists(done_interventions),
    )

    # TODO: this pilot endpoint will be replaced by the registry-driven map feed later.
    bbox_param = request.GET.get("bbox")
    if bbox_param:
        try:
            min_lon, min_lat, max_lon, max_lat = [float(part) for part in bbox_param.split(",")]
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid bbox parameter"}, status=400)
        if min_lon > max_lon:
            min_lon, max_lon = max_lon, min_lon
        if min_lat > max_lat:
            min_lat, max_lat = max_lat, min_lat
        qs = qs.filter(
            longitude__gte=min_lon,
            longitude__lte=max_lon,
            latitude__gte=min_lat,
            latitude__lte=max_lat,
        )

    features = []
    for wr in qs:
        label = wr.display_label

        intervention_stage = "none"
        if getattr(wr, "has_approved_intervention", False):
            intervention_stage = "approved"
        elif getattr(wr, "has_done_intervention", False):
            intervention_stage = "done"

        shrub_width_value = None
        if wr.vegetation_type == WorkRecord.VegetationType.HEDGE:
            shrub_width_value = to_float(getattr(wr, "shrub_width_m", None))

        features.append(
            {
                "type": "Feature",
                "id": wr.id,
                "geometry": {
                    "type": "Point",
                    "coordinates": [wr.longitude, wr.latitude],
                },
                "properties": {
                    "id": wr.id,
                    "label": label,
                    "map_label": wr.map_label,
                    "vegetation_type": wr.vegetation_type,
                    "crown_width_m": to_float(getattr(wr, "crown_width_m", None)),
                    "hedge_line": wr.hedge_line,
                    "shrub_width_m": shrub_width_value,
                    "intervention_stage": intervention_stage,
                },
            }
        )

    return JsonResponse(
        {
            "type": "FeatureCollection",
            "features": features,
        },
        safe=False,
    )


@login_required
def map_gl_pilot(request):
    """Temporary pilot page to verify MapLibre GL rendering."""
    project_param = request.GET.get("project")
    if project_param:
        try:
            project_id = int(project_param)
        except (TypeError, ValueError):
            project_id = None
        if project_id:
            request.session["active_project_id"] = project_id
            request.session["active_project_return"] = request.get_full_path()
    context = _build_map_mapui_context(request)
    return render(request, "tracker/map_gl_pilot.html", context)


@login_required
def map_project_redirect(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if not user_can_view_project(request.user, project.pk):
        return redirect("work_record_list")

    passthrough_params = {}
    for key in ("focus", "lat", "lon", "z"):
        value = request.GET.get(key)
        if value:
            passthrough_params[key] = value


    coords_qs = project.trees.filter(
        latitude__isnull=False,
        longitude__isnull=False,
    )
    coords_count = coords_qs.count()
    base_params = {"project": project.pk, **passthrough_params}
    if coords_count == 0:
        messages.warning(request, "Projekt nemá žádné stromy se souřadnicemi.")
        target = f"{reverse('map_gl_pilot')}?{urlencode(base_params)}"
        return redirect(target)
    if coords_count == 1:
        single = coords_qs.values("latitude", "longitude").first()
        if single:
            lat = single["latitude"]
            lon = single["longitude"]
            params = dict(base_params)
            params.setdefault("lat", lat)
            params.setdefault("lon", lon)
            params.setdefault("z", 18)
            return redirect(f"{reverse('map_gl_pilot')}?{urlencode(params)}")

    coords = coords_qs.aggregate(
        min_lat=Min("latitude"),
        max_lat=Max("latitude"),
        min_lon=Min("longitude"),
        max_lon=Max("longitude"),
    )

    target = reverse("map_gl_pilot")
    if all(value is not None for value in coords.values()):
        bbox = f"{coords['min_lon']},{coords['min_lat']},{coords['max_lon']},{coords['max_lat']}"
        params = dict(base_params)
        params.setdefault("bbox", bbox)
        return redirect(f"{target}?{urlencode(params)}")
    return redirect(f"{target}?{urlencode(base_params)}")


@login_required
@require_http_methods(["POST"])
def project_tree_add(request, project_pk, workrecord_pk):
    project = get_object_or_404(Project, pk=project_pk)
    if not user_is_foreman(request.user, project.pk):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)
    work_record = get_object_or_404(WorkRecord, pk=workrecord_pk)
    project.trees.add(work_record)
    if work_record.project_id is None:
        work_record.project = project
        work_record.save(update_fields=["project"])
    return JsonResponse({"ok": True})


@login_required
def pmtiles_range_serve(request, path):
    """
    Serve PMTiles with HTTP Range support for MapLibre/pmtiles.js.
    """
    static_path = finders.find(f"tiles/{path}")
    if not static_path or not os.path.isfile(static_path):
        return HttpResponse(status=404)

    file_size = os.path.getsize(static_path)
    range_header = request.headers.get("Range") or request.META.get("HTTP_RANGE")
    content_type = "application/octet-stream"

    if range_header and range_header.startswith("bytes="):
        try:
            range_value = range_header.split("=", 1)[1]
            start_str, end_str = (range_value.split("-", 1) + [""])[:2]
            start = int(start_str) if start_str else 0
            end = int(end_str) if end_str else file_size - 1
        except (ValueError, IndexError):
            return HttpResponse(status=416)

        if start < 0 or end < start or end >= file_size:
            return HttpResponse(status=416)

        def range_stream():
            with open(static_path, "rb") as f:
                f.seek(start)
                remaining = end - start + 1
                chunk_size = 8192
                while remaining > 0:
                    chunk = f.read(min(chunk_size, remaining))
                    if not chunk:
                        break
                    remaining -= len(chunk)
                    yield chunk

        resp = StreamingHttpResponse(range_stream(), status=206, content_type=content_type)
        resp["Content-Length"] = str(end - start + 1)
        resp["Content-Range"] = f"bytes {start}-{end}/{file_size}"
        resp["Accept-Ranges"] = "bytes"
        return resp

    # Full response
    response = FileResponse(open(static_path, "rb"), content_type=content_type)
    response["Content-Length"] = str(file_size)
    response["Accept-Ranges"] = "bytes"
    return response


@login_required
@require_GET
def workrecord_detail_api(request, pk):
    work_record = (
        WorkRecord.objects
        .filter(pk=pk)
        .select_related("project")
        .prefetch_related(
            Prefetch("photos", queryset=PhotoDocumentation.objects.order_by("-id"))
        )
        .first()
    )
    if not work_record:
        return JsonResponse({"status": "error", "msg": "WorkRecord nenalezen"}, status=404)

    if work_record.project_id and not user_can_view_project(request.user, work_record.project_id):
        return JsonResponse({"status": "error", "msg": "Nemáte oprávnění."}, status=403)

    photos = []
    for photo in work_record.photos.all():
        if not photo.photo:
            continue
        full_url = photo.photo.url
        thumb_url = get_photo_thumbnail(photo)
        photos.append({
            "id": photo.id,
            "thumb": thumb_url or full_url,
            "full": full_url,
            "description": photo.description or "",
        })

    if work_record.vegetation_type in (
        WorkRecord.VegetationType.SHRUB,
        WorkRecord.VegetationType.HEDGE,
    ):
        has_assessment = work_record.shrub_assessments.exists()
    else:
        has_assessment = work_record.assessments.exists()
    can_edit = can_edit_project(request.user, work_record.project)

    return JsonResponse({
        "status": "ok",
        "record": {
            "id": work_record.id,
            "title": work_record.title or "",
            "external_tree_id": work_record.external_tree_id or "",
            "passport_code": work_record.passport_code or "",
            "display_label": work_record.display_label,
            "taxon": work_record.taxon or "",
            "vegetation_type": work_record.vegetation_type,
            "project": work_record.project.name if work_record.project else "",
            "project_id": work_record.project_id,
            "lat": work_record.latitude,
            "lon": work_record.longitude,
            "can_edit": can_edit,
            "has_assessment": has_assessment,
            "has_photos": bool(photos),
            "photos": photos,
        },
    })

@login_required
@require_GET
def gbif_taxon_suggest(request):
    q = (request.GET.get("q") or "").strip()
    if len(q) < 2:
        return JsonResponse({"results": []})

    qs = (
        Species.objects
        .filter(Q(latin_name__icontains=q) | Q(czech_name__icontains=q))
        .order_by("latin_name")[:20]
    )

    results = []
    for sp in qs:
        latin = sp.latin_name
        czech = sp.czech_name or ""
        if czech:
            display = f"{czech} ({latin})"
        else:
            display = latin

        results.append(
            {
                "gbif_key": None,
                "scientific_name": latin,
                "vernacular_name": czech,
                "display": display,
            }
        )

    return JsonResponse({"results": results})

@login_required
def save_coordinates(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "msg": "Invalid request"}, status=405)

    record_id = request.POST.get("record_id")
    lat_str = request.POST.get("latitude")
    lon_str = request.POST.get("longitude")

    try:
        record = WorkRecord.objects.get(id=record_id)
    except WorkRecord.DoesNotExist:
        return JsonResponse({"status": "error", "msg": "Record not found"}, status=404)

    if record.project_id and not user_can_view_project(request.user, record.project_id):
        return JsonResponse({"status": "error", "msg": "Forbidden"}, status=403)

    try:
        lat = float(lat_str)
        lon = float(lon_str)
    except (TypeError, ValueError):
        return JsonResponse({"status": "error", "msg": "Invalid coordinates"}, status=400)

    record.latitude = lat
    record.longitude = lon
    record.save(update_fields=["latitude", "longitude"])
    return JsonResponse({"status": "ok"})


@login_required
@require_http_methods(["POST"])
def workrecord_set_location(request, pk):
    work_record = get_object_or_404(WorkRecord, pk=pk)
    if not can_edit_project(request.user, work_record.project):
        return JsonResponse({"error": "Forbidden"}, status=403)

    payload = {}
    if request.POST:
        payload = request.POST
    elif request.body:
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid payload"}, status=400)

    def parse_float(value):
        if value in (None, ""):
            return None
        try:
            return float(str(value).replace(",", "."))
        except (TypeError, ValueError):
            return None

    lat = parse_float(payload.get("latitude") or payload.get("lat"))
    lon = parse_float(payload.get("longitude") or payload.get("lon"))
    if lat is None or lon is None:
        return JsonResponse({"error": "Invalid coordinates"}, status=400)

    work_record.latitude = lat
    work_record.longitude = lon
    work_record.save(update_fields=["latitude", "longitude"])
    return JsonResponse({"id": work_record.pk, "latitude": lat, "longitude": lon})


@login_required
def map_upload_photo(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "msg": "Invalid request"}, status=405)

    record_id = request.POST.get("record_id")
    comment = request.POST.get("comment", "").strip()
    photo_file = request.FILES.get("photo")

    if not record_id or not photo_file:
        return JsonResponse({"status": "error", "msg": "Chybí data"}, status=400)

    try:
        record = WorkRecord.objects.get(id=record_id)
    except WorkRecord.DoesNotExist:
        return JsonResponse({"status": "error", "msg": "Úkon nenalezen"}, status=404)

    if record.project_id and not user_can_view_project(request.user, record.project_id):
        return JsonResponse({"status": "error", "msg": "Nemáš oprávnění"}, status=403)

    photo_doc = PhotoDocumentation.objects.create(
        work_record=record,
        photo=photo_file,
        description=comment
    )

    thumb_url = get_photo_thumbnail(photo_doc)
    return JsonResponse({
        "status": "ok",
        "photo": {
            "id": photo_doc.id,
            "thumb": thumb_url or (photo_doc.photo.url if photo_doc.photo else ""),
            "full": photo_doc.photo.url if photo_doc.photo else "",
            "description": comment,
        }
    })


@login_required
def map_create_work_record(request):
    """
    Vytvoří nový úkon přímo z mapy (AJAX, bez přesměrování).
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "msg": "Invalid request"}, status=405)

    external_tree_id = (request.POST.get("external_tree_id") or "").strip()
    taxon_value = (request.POST.get("taxon") or "").strip()
    taxon_czech_value = (request.POST.get("taxon_czech") or "").strip()
    taxon_latin_value = (request.POST.get("taxon_latin") or "").strip()
    gbif_key_raw = (request.POST.get("taxon_gbif_key") or "").strip()
    vegetation_type_raw = (request.POST.get("vegetation_type") or "").strip().upper()
    project_id = request.POST.get("project_id") or request.POST.get("project") or request.GET.get("project") or None
    date_str = (request.POST.get("date") or "").strip()
    lat_str = request.POST.get("latitude")
    lon_str = request.POST.get("longitude")
    hedge_line_raw = request.POST.get("hedge_line")

    # projekt (volitelný)
    project = None
    if project_id and project_id != "none":
        try:
            project = Project.objects.get(pk=project_id)
        except Project.DoesNotExist:
            return JsonResponse({"status": "error", "msg": "Projekt nenalezen."}, status=404)
        if not user_can_view_project(request.user, project.pk):
            return JsonResponse({"status": "error", "msg": "Nemáš oprávnění."}, status=403)

    # datum – pokud nepřijde, použijeme dnešek
    from django.utils import timezone

    if date_str:
        parsed = parse_date(date_str)
        record_date = parsed or timezone.localdate()
    else:
        record_date = timezone.localdate()

    taxon_gbif_key = None
    if gbif_key_raw:
        try:
            taxon_gbif_key = int(gbif_key_raw)
        except ValueError:
            taxon_gbif_key = None

    valid_types = {choice[0] for choice in WorkRecord.VegetationType.choices}
    vegetation_type = (
        vegetation_type_raw
        if vegetation_type_raw in valid_types
        else WorkRecord.VegetationType.TREE
    )

    def parse_hedge_line(raw_value):
        if not raw_value:
            return None
        data = raw_value
        if isinstance(raw_value, str):
            try:
                data = json.loads(raw_value)
            except json.JSONDecodeError:
                return None
        if not isinstance(data, dict):
            return None
        if data.get("type") != "LineString":
            return None
        coords = data.get("coordinates")
        if not isinstance(coords, list) or len(coords) < 2:
            return None
        normalized = []
        for point in coords:
            if not isinstance(point, (list, tuple)) or len(point) < 2:
                return None
            try:
                lon = float(point[0])
                lat = float(point[1])
            except (TypeError, ValueError):
                return None
            normalized.append([lon, lat])
        return {"type": "LineString", "coordinates": normalized}

    hedge_line = None
    if vegetation_type == WorkRecord.VegetationType.HEDGE:
        hedge_line = parse_hedge_line(hedge_line_raw)
        if not hedge_line:
            return JsonResponse(
                {"status": "error", "msg": "Chybné GeoJSON LineString pro živý plot."},
                status=400,
            )
        coords = hedge_line.get("coordinates", [])
        count = len(coords)
        if count == 0:
            return JsonResponse(
                {"status": "error", "msg": "Živý plot musí mít alespoň 2 body."},
                status=400,
            )
        lat = sum(point[1] for point in coords) / count
        lon = sum(point[0] for point in coords) / count
    else:
        # souřadnice jsou pro mapu povinné
        try:
            lat = float(lat_str)
            lon = float(lon_str)
        except (TypeError, ValueError):
            return JsonResponse({"status": "error", "msg": "Chybné souřadnice."}, status=400)

    work_record = WorkRecord(
        project=project,
        external_tree_id=external_tree_id or None,
        vegetation_type=vegetation_type,
        taxon=taxon_value or "",
        taxon_czech=taxon_czech_value,
        taxon_latin=taxon_latin_value,
        taxon_gbif_key=taxon_gbif_key,
        latitude=lat,
        longitude=lon,
        hedge_line=hedge_line,
        date=record_date,
    )
    work_record.save()
    work_record.assign_passport_identifiers()
    work_record.sync_title_from_identifiers()
    work_record.save(
        update_fields=[
            "title",
            "external_tree_id",
            "taxon",
            "taxon_czech",
            "taxon_latin",
            "taxon_gbif_key",
            "project",
            "latitude",
            "longitude",
            "hedge_line",
            "date",
        ]
    )

    if project:
        project.trees.add(work_record)
        if settings.DEBUG:
            print(f"[map_create_work_record] project_id={project.pk} work_record_id={work_record.pk}")

    return JsonResponse({
        "status": "ok",
        "record": {
            "id": work_record.id,
            "title": work_record.title or "",
            "external_tree_id": work_record.external_tree_id or "",
            "passport_code": work_record.passport_code or "",
            "display_label": work_record.display_label,
            "map_label": work_record.map_label,
            "vegetation_type": work_record.vegetation_type,
            "taxon": work_record.taxon or "",
            "taxon_czech": work_record.taxon_czech or "",
            "taxon_latin": work_record.taxon_latin or "",
            "taxon_gbif_key": work_record.taxon_gbif_key,
            "project": work_record.project.name if work_record.project else "",
            "project_id": work_record.project_id,
            "lat": work_record.latitude,
            "lon": work_record.longitude,
            "hedge_line": work_record.hedge_line,
        },
    })


@login_required
@require_http_methods(["GET", "POST"])
def workrecord_assessment_api(request, pk):
    """
    JSON API for reading TreeAssessments tied to a WorkRecord.
    GET: return the latest assessment values (or nulls if none exist).
    POST: append a new assessment version from JSON payload.
    """
    try:
        work_record = WorkRecord.objects.get(pk=pk)
    except WorkRecord.DoesNotExist:
        return JsonResponse({"error": "WorkRecord not found"}, status=404)

    if request.method == "GET":
        assessment = work_record.latest_assessment
        data = {
            "work_record_id": work_record.pk,
            "dbh_cm": assessment.dbh_cm if assessment else None,
            "stem_circumference_cm": assessment.stem_circumference_cm if assessment else None,
            "stem_diameters_cm_list": assessment.stem_diameters_cm_list if assessment else "",
            "stem_circumferences_cm_list": assessment.stem_circumferences_cm_list if assessment else "",
            "height_m": assessment.height_m if assessment else None,
            "crown_width_m": str(assessment.crown_width_m) if assessment and assessment.crown_width_m is not None else None,
            "crown_area_m2": str(assessment.crown_area_m2) if assessment and assessment.crown_area_m2 is not None else None,
            "physiological_age": assessment.physiological_age if assessment else None,
            "vitality": assessment.vitality if assessment else None,
            "health_state": assessment.health_state if assessment else None,
            "stability": assessment.stability if assessment else None,
            "mistletoe_level": assessment.mistletoe_level if assessment else None,
            "perspective": assessment.perspective if assessment else None,
            "assessed_at": assessment.assessed_at.isoformat() if assessment and assessment.assessed_at else None,
        }
        return JsonResponse(data)

    # POST
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Invalid JSON")

    def parse_int(value, min_val, max_val):
        if value in (None, ""):
            return None
        try:
            iv = int(value)
        except (TypeError, ValueError):
            return None
        if iv < min_val or iv > max_val:
            return None
        return iv

    def parse_float(value):
        if value in (None, ""):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def parse_decimal(value):
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return None

    def parse_cm_int(value):
        if value in (None, ""):
            return None
        try:
            num = float(value)
        except (TypeError, ValueError):
            return None
        rounded = int(round(num))
        if rounded <= 0:
            return None
        return rounded

    def parse_cm_list(value):
        if value in (None, ""):
            return []
        if isinstance(value, (list, tuple)):
            raw_items = value
        else:
            raw_items = str(value).split(",")
        items = []
        for item in raw_items:
            parsed = parse_cm_int(item)
            if parsed is not None:
                items.append(parsed)
        return items

    def normalize_cm_lists(diameters, circumferences):
        length = max(len(diameters), len(circumferences))
        norm_diameters = []
        norm_circumferences = []
        for i in range(length):
            d_val = diameters[i] if i < len(diameters) else None
            c_val = circumferences[i] if i < len(circumferences) else None
            if d_val is None and c_val is None:
                continue
            if d_val is None and c_val is not None:
                d_val = int(round(c_val / math.pi))
            if c_val is None and d_val is not None:
                c_val = int(round(d_val * math.pi))
            if not d_val or not c_val:
                continue
            norm_diameters.append(d_val)
            norm_circumferences.append(c_val)
        return norm_diameters, norm_circumferences

    dbh_cm = parse_cm_int(payload.get("dbh_cm"))
    stem_circumference_cm = parse_cm_int(payload.get("stem_circumference_cm"))
    stem_diameters_cm_list = parse_cm_list(payload.get("stem_diameters_cm_list"))
    stem_circumferences_cm_list = parse_cm_list(payload.get("stem_circumferences_cm_list"))
    height_m = parse_float(payload.get("height_m"))
    crown_width_m = parse_decimal(payload.get("crown_width_m"))
    physiological_age = parse_int(payload.get("physiological_age"), 1, 5)
    vitality = parse_int(payload.get("vitality"), 1, 5)
    health_state = parse_int(payload.get("health_state"), 1, 5)
    stability = parse_int(payload.get("stability"), 1, 5)
    raw_mistletoe = payload.get("mistletoe_level")
    if raw_mistletoe in (None, "", 0, "0"):
        mistletoe_level = None
    else:
        mistletoe_level = parse_int(raw_mistletoe, 1, 5)
    perspective = payload.get("perspective") or None
    if perspective not in (None, "", "a", "b", "c"):
        perspective = None
    norm_diameters, norm_circumferences = normalize_cm_lists(
        stem_diameters_cm_list, stem_circumferences_cm_list
    )
    if not norm_diameters and (dbh_cm is not None or stem_circumference_cm is not None):
        if dbh_cm is None and stem_circumference_cm is not None:
            dbh_cm = int(round(stem_circumference_cm / math.pi))
        elif stem_circumference_cm is None and dbh_cm is not None:
            stem_circumference_cm = int(round(dbh_cm * math.pi))
        if dbh_cm is not None and stem_circumference_cm is not None:
            norm_diameters = [dbh_cm]
            norm_circumferences = [stem_circumference_cm]
    if norm_diameters:
        max_idx = max(range(len(norm_diameters)), key=lambda idx: norm_diameters[idx])
        dbh_cm = norm_diameters[max_idx]
        stem_circumference_cm = norm_circumferences[max_idx]
    elif dbh_cm is None and stem_circumference_cm is not None:
        dbh_cm = int(round(stem_circumference_cm / math.pi))
    elif stem_circumference_cm is None and dbh_cm is not None:
        stem_circumference_cm = int(round(dbh_cm * math.pi))

    assessment = TreeAssessment.objects.create(
        work_record=work_record,
        assessed_at=date.today(),
        dbh_cm=dbh_cm,
        stem_circumference_cm=stem_circumference_cm,
        stem_diameters_cm_list=",".join(str(val) for val in norm_diameters),
        stem_circumferences_cm_list=",".join(str(val) for val in norm_circumferences),
        height_m=height_m,
        crown_width_m=crown_width_m,
        physiological_age=physiological_age,
        vitality=vitality,
        health_state=health_state,
        stability=stability,
        mistletoe_level=mistletoe_level,
        perspective=perspective,
    )

    return JsonResponse({
        "status": "ok",
        "id": assessment.pk,
        "work_record_id": work_record.pk,
        "dbh_cm": assessment.dbh_cm,
        "stem_circumference_cm": assessment.stem_circumference_cm,
        "stem_diameters_cm_list": assessment.stem_diameters_cm_list,
        "stem_circumferences_cm_list": assessment.stem_circumferences_cm_list,
        "height_m": assessment.height_m,
        "crown_width_m": str(assessment.crown_width_m) if assessment.crown_width_m is not None else None,
        "crown_area_m2": str(assessment.crown_area_m2) if assessment.crown_area_m2 is not None else None,
        "physiological_age": assessment.physiological_age,
        "vitality": assessment.vitality,
        "health_state": assessment.health_state,
        "stability": assessment.stability,
        "mistletoe_level": assessment.mistletoe_level,
        "perspective": assessment.perspective,
        "assessed_at": assessment.assessed_at.isoformat() if assessment.assessed_at else None,
    })


@login_required
@require_http_methods(["GET", "POST"])
def workrecord_shrub_assessment_api(request, pk):
    """
    JSON API for reading ShrubAssessments tied to a WorkRecord.
    GET: return the latest assessment values (or nulls if none exist).
    POST: append a new assessment version from JSON payload.
    """
    try:
        work_record = WorkRecord.objects.get(pk=pk)
    except WorkRecord.DoesNotExist:
        return JsonResponse({"error": "WorkRecord not found"}, status=404)

    if work_record.vegetation_type not in (
        WorkRecord.VegetationType.SHRUB,
        WorkRecord.VegetationType.HEDGE,
    ):
        return JsonResponse(
            {"error": "WorkRecord is not shrub/hedge"},
            status=400,
        )

    if request.method == "GET":
        assessment = work_record.latest_shrub_assessment
        data = {
            "work_record_id": work_record.pk,
            "height_m": assessment.height_m if assessment else None,
            "width_m": assessment.width_m if assessment else None,
            "vitality": assessment.vitality if assessment else None,
            "note": assessment.note if assessment else "",
            "assessed_at": assessment.assessed_at.isoformat() if assessment and assessment.assessed_at else None,
        }
        return JsonResponse(data)

    # POST
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Invalid JSON")

    def parse_int(value, min_val, max_val):
        if value in (None, ""):
            return None
        try:
            iv = int(value)
        except (TypeError, ValueError):
            return None
        if iv < min_val or iv > max_val:
            return None
        return iv

    def parse_float(value):
        if value in (None, ""):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    height_m = parse_float(payload.get("height_m"))
    width_m = parse_float(payload.get("width_m"))
    vitality = parse_int(payload.get("vitality"), 1, 5)
    note = (payload.get("note") or "").strip()

    # Always append a new record to preserve assessment history.
    assessment = ShrubAssessment.objects.create(
        work_record=work_record,
        assessed_at=date.today(),
        height_m=height_m,
        width_m=width_m,
        vitality=vitality,
        note=note,
    )

    return JsonResponse({
        "status": "ok",
        "id": assessment.pk,
        "work_record_id": work_record.pk,
        "height_m": assessment.height_m,
        "width_m": assessment.width_m,
        "vitality": assessment.vitality,
        "note": assessment.note,
        "assessed_at": assessment.assessed_at.isoformat() if assessment.assessed_at else None,
    })


@login_required
def bulk_handover_interventions(request, pk):
    project = get_object_or_404(Project, pk=pk)

    if not user_can_view_project(request.user, project.pk):
        return redirect('work_record_list')

    if request.method != "POST":
        return redirect("project_detail", pk=pk)

    selected_ids = request.POST.getlist("selected_records")
    if not selected_ids:
        messages.warning(request, "Vyberte prosím alespoň jeden strom / úkon.")
        return redirect("project_detail", pk=pk)

    work_records = project.trees.filter(id__in=selected_ids)
    interventions_qs = TreeIntervention.objects.filter(
        tree__in=work_records,
        status__in=["proposed"],
    )
    interventions = list(interventions_qs)

    if not interventions:
        messages.info(request, "Nebyl nalezen žádný zásah vhodný k předání ke kontrole.")
        return redirect("project_detail", pk=pk)

    for intervention in interventions:
        if not can_transition_intervention(request.user, intervention, "done_pending_owner"):
            return HttpResponse(status=403)
        intervention.status = "done_pending_owner"
        if getattr(intervention, "handed_over_for_check_at", None) is None:
            intervention.handed_over_for_check_at = timezone.now()
        intervention.save(update_fields=["status", "handed_over_for_check_at"])
    messages.success(
        request,
        f"Předáno ke kontrole {len(interventions)} zásahů.",
    )
    return redirect("project_detail", pk=pk)


@login_required
def bulk_complete_interventions(request, pk):
    project = get_object_or_404(Project, pk=pk)

    if not user_can_view_project(request.user, project.pk):
        return redirect('work_record_list')

    if request.method != "POST":
        return redirect("project_detail", pk=pk)

    selected_ids = request.POST.getlist("selected_records")
    if not selected_ids:
        messages.warning(request, "Vyberte prosím alespoň jeden strom / úkon.")
        return redirect("project_detail", pk=pk)

    work_records = project.trees.filter(id__in=selected_ids)
    interventions_qs = TreeIntervention.objects.filter(
        tree__in=work_records,
        status__in=["done_pending_owner"],
    )
    interventions = list(interventions_qs)

    if not interventions:
        messages.info(request, "Nebyl nalezen žádný zásah k označení jako dokončený.")
        return redirect("project_detail", pk=pk)

    for intervention in interventions:
        if not can_transition_intervention(request.user, intervention, "completed"):
            return HttpResponse(status=403)
        intervention.status = "completed"
        intervention.save(update_fields=["status"])
    messages.success(
        request,
        f"Označeno {len(interventions)} zásahů jako dokončené.",
    )
    return redirect("project_detail", pk=pk)
