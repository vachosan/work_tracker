import re
import unicodedata

from django.core.management.base import BaseCommand

from tracker.models import PriceListItem, PriceListVersion


ITEM_CODE_RE = re.compile(r"^ZE41[a-z]+$", re.IGNORECASE)


def _normalize(text: str) -> str:
    text = text or ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.lower().strip()


def _find_sheet(workbook):
    for name in workbook.sheetnames:
        norm = _normalize(name)
        if "zelen" in norm and ("mimo" in norm or "les" in norm):
            return workbook[name]
    return None


def _parse_band(label: str):
    if not label:
        return None, None
    norm = _normalize(label)
    match = re.search(r"do\s*(\d+)", norm)
    if match:
        return 0, int(match.group(1))
    match = re.search(r"(\d+)\s*[-–]\s*(\d+)", norm)
    if match:
        return int(match.group(1)), int(match.group(2))
    if "vice nez" in norm or "nad" in norm:
        match = re.search(r"(?:vice nez|nad)\s*(\d+)", norm)
        if match:
            return int(match.group(1)) + 1, None
    return None, None


def _parse_operation_type(label: str):
    norm = _normalize(label)
    if "kombinace" in norm:
        return "kombinace", True
    if norm.startswith("zdravotni rez"):
        return "zdravotni", False
    if norm.startswith("bezpecnostni rez"):
        return "bezpecnostni", False
    if norm.startswith("lokalni redukce"):
        return "lokalni", False
    if "obvodova redukce" in norm:
        return "obvodova", False
    return "jine", False


def _is_memorial_or_special(label: str) -> bool:
    norm = _normalize(label)
    return "pamatne" in norm or "vyjimecne" in norm


def _coerce_price(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(round(value))
    try:
        cleaned = str(value).strip().replace(" ", "").replace(",", ".")
        return int(round(float(cleaned)))
    except (TypeError, ValueError):
        return None


class Command(BaseCommand):
    help = "Import NOO 2026 price list items (ZE41*) from Excel."

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            required=True,
            help="Path to NOO 2026 Excel file.",
        )

    def handle(self, *args, **options):
        path = options["path"]
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise SystemExit(
                "openpyxl is required for import_noo_2026. Please install it."
            ) from exc

        workbook = load_workbook(path, data_only=True)
        sheet = _find_sheet(workbook)
        if not sheet:
            raise SystemExit("Sheet not found: expected Zeleň rostoucí mimo les.")

        version, _ = PriceListVersion.objects.get_or_create(
            code="NOO_2026",
            defaults={"label": "NOO 2026"},
        )
        if version.label != "NOO 2026":
            version.label = "NOO 2026"
            version.save(update_fields=["label"])

        imported = 0
        with_band = 0
        combo_count = 0
        examples = []

        for row in sheet.iter_rows(values_only=True):
            if not row or len(row) < 2:
                continue
            raw_code = row[0]
            raw_label = row[1]
            if not raw_code or not raw_label:
                continue
            item_code = str(raw_code).strip()
            if not ITEM_CODE_RE.match(item_code):
                continue
            label = str(raw_label).strip()
            price = None
            for cell in reversed(row):
                price = _coerce_price(cell)
                if price is not None:
                    break
            if price is None:
                continue
            unit = None
            if len(row) >= 3 and row[2]:
                unit = str(row[2]).strip()
            if not unit:
                unit = "ks"

            band_min, band_max = _parse_band(label)
            if band_min is not None or band_max is not None:
                with_band += 1

            operation_type, is_combo = _parse_operation_type(label)
            if is_combo:
                combo_count += 1

            is_memorial = _is_memorial_or_special(label)

            defaults = {
                "activity_code": "ZE41",
                "label": label,
                "unit": unit,
                "price_czk": price,
                "band_min_m2": band_min,
                "band_max_m2": band_max,
                "operation_type": operation_type,
                "is_combo": is_combo,
                "is_memorial_or_special": is_memorial,
                "metadata": {
                    "operation_type": operation_type,
                    "is_combo": is_combo,
                    "is_memorial_or_special": is_memorial,
                    "raw_label": label,
                },
            }

            PriceListItem.objects.update_or_create(
                version=version,
                item_code=item_code,
                defaults=defaults,
            )
            imported += 1
            if len(examples) < 5:
                band_label = (
                    f"{band_min or ''}-{band_max or ''}".strip("-") if band_min or band_max else "-"
                )
                examples.append(f"{item_code} | {operation_type} | {band_label} | {price}")

        self.stdout.write(self.style.SUCCESS(f"Imported ZE41 items: {imported}"))
        self.stdout.write(self.style.SUCCESS(f"With band: {with_band}"))
        self.stdout.write(self.style.SUCCESS(f"Combos: {combo_count}"))
        if examples:
            self.stdout.write("Examples:")
            for example in examples:
                self.stdout.write(f"- {example}")
